# coding=utf-8
from pykeyboard import PyKeyboard
from pymouse import PyMouse
import time
import pyHook
import pythoncom
import xlrd
import xlwt
import pyperclip
from pynput import mouse, keyboard
import threading
import sys
from openpyxl import Workbook, load_workbook
import os


def copy():
    k.press_key(k.control_key)
    k.tap_key("c")  # 改小写！！！！ 大写的话由于单进程会触发shift键 ctrl键就失效了
    k.release_key(k.control_key)


def getCopy(maxTime=2):
    # maxTime = 3  # 3秒复制 调用copy() 不管结果对错
    while (maxTime > 0):
        maxTime = maxTime - 0.5
        time.sleep(0.5)
        # print('doing')
        copy()

    result = pyperclip.paste()
    return result


def tapkey(key, count=1):
    for i in range(0, count):
        k.tap_key(key)
        time.sleep(0.2)


def Do():
    global start
    global curExcelUrls
    global end
    if start:
        update_workbook = xlrd.open_workbook(updateExcelUrl)
        u_table = update_workbook.sheets()[5]#最后一个sheet
        u_rowCount = u_table.nrows
        u_colCount = u_table.ncols

        wb = load_workbook(filename=updateExcelUrl)
        load_worksheet = wb.active
        load_worksheet = wb['update']


        curSaveAddRow=1
        addLists=[]

        # 主代码---------------
        for curExcelUrl in curExcelUrls:
            workbook = xlrd.open_workbook(curDirUrl+'\\'+curExcelUrl)
            print('开始：'+curExcelUrl)
            table = workbook.sheets()[0]
            rowCount = table.nrows
            colCount = table.ncols

            # 中间的标题--------------------------------①
            # u_rowCount = u_table.nrows
            #
            # load_worksheet.cell(row=u_rowCount + curSaveAddRow, column=2, value=' ')
            # curSaveAddRow=curSaveAddRow + 1
            # load_worksheet.cell(row=u_rowCount + curSaveAddRow, column=2, value=curExcelUrl)
            # curSaveAddRow=curSaveAddRow + 1
            # load_worksheet.cell(row=u_rowCount + curSaveAddRow, column=2, value=' ')
            # curSaveAddRow=curSaveAddRow + 1

            # -----------------①


            for i in range(rowCount):
                name=str(table.cell_value(i,0))#名称
                features=str(table.cell_value(i,1))#特征
                unit=str(table.cell_value(i,2))#单位
                value=str(table.cell_value(i,3))#值

                isExist=False
                #之后每次更新用这个---------------②
                for u_row in range(u_rowCount):
                    u_name = str(u_table.cell_value(u_row, 0))  # 需更新的名称
                    u_features = str(u_table.cell_value(u_row, 1))  # 需更新的特征
                    u_value=str(u_table.cell_value(u_row,2)) # 需更新的值
                    u_unit = str(u_table.cell_value(u_row, 3)) #单位
                    if(u_name==name and u_features==features and u_value!=value):
                        print('更新：')
                        print(name)
                        print(features)
                        print(value)

                        load_worksheet.cell(row=u_row+1, column=3, value=value)#改值

                    if(isExist == False):
                        isExist=u_name == name and u_features == features

                        if(isExist is False):
                            # 只改了特征，但价格，名称相同的不算新增-----
                            isExist=u_name ==name and value==u_value and u_features!=features
                            # if(isExist):
                            #     print('改了特征，但价格相同-------')
                            #     print('广材中：')
                            #     print(name)
                            #     print(features)
                            #     print(value)
                            #     print('xing中：')
                            #     print(u_name)
                            #     print(u_features)
                            #     print(u_value)

                # 之后每次更新用这个---------------②


            # 空白填入-------------------①
            # u_rowCount = u_table.nrows
            # u_colCount = u_table.ncols
            # load_worksheet.cell(row=u_rowCount+curSaveAddRow, column=1, value=name)
            # load_worksheet.cell(row=u_rowCount+curSaveAddRow, column=2, value=features)
            # load_worksheet.cell(row=u_rowCount+curSaveAddRow, column=3, value=value)
            # load_worksheet.cell(row=u_rowCount+curSaveAddRow, column=4, value=unit)
            # curSaveAddRow=curSaveAddRow+1
            # ---------------------------------①

        # 更新添加新项目---------------②
                if(not isExist):
                    print('新添加：')
                    print(name)
                    print(features)
                    l=[]
                    l.append(name)
                    l.append(features)
                    l.append(value)
                    l.append(unit)

                    addlist_isExist=False
                    for add in addLists:
                        _n = l[0]
                        _f = l[1]
                        if(_n==name and _f==features):
                            addlist_isExist=True
                    if(addlist_isExist==False):
                        addLists.append(l)

        if(len(addLists)>0):
            u_rowCount = u_table.nrows
            load_worksheet.cell(row=u_rowCount + curSaveAddRow, column=2, value=' ')
            curSaveAddRow = curSaveAddRow + 1
            load_worksheet.cell(row=u_rowCount + curSaveAddRow, column=2, value='----新添加----')
            curSaveAddRow = curSaveAddRow + 1
            load_worksheet.cell(row=u_rowCount + curSaveAddRow, column=2, value=' ')
            curSaveAddRow = curSaveAddRow + 1
        for l in addLists:
            list_name=l[0]
            list_features=l[1]
            list_value=l[2]
            list_unit=l[3]

            u_rowCount = u_table.nrows
            u_colCount = u_table.ncols
            load_worksheet.cell(row=u_rowCount + curSaveAddRow, column=1, value=list_name)
            load_worksheet.cell(row=u_rowCount + curSaveAddRow, column=2, value=list_features)
            load_worksheet.cell(row=u_rowCount + curSaveAddRow, column=3, value=list_value)
            load_worksheet.cell(row=u_rowCount + curSaveAddRow, column=4, value=list_unit)
            curSaveAddRow = curSaveAddRow + 1
        # 更新添加新项目---------------②




        print('存储中……')
        wb.save(updateExcelUrl)
        print('存储完毕!!!')
    start=False



def saveToExcel(name, type, unit, value):
    # saveworkbook = xlrd.open_workbook(saveExcelUrl)
    # wb = excel_copy(saveworkbook)  # 利用xlutils.copy下的copy函数复制
    wb = load_workbook(filename=saveExcelUrl)
    worksheet = wb.active
    worksheet = wb['Sheet1']
    global rowMaxCount
    # print(rowMaxCount)
    worksheet.cell(row=rowMaxCount + 1, column=1, value=name)
    worksheet.cell(row=rowMaxCount + 1, column=2, value=type)
    worksheet.cell(row=rowMaxCount + 1, column=3, value=unit)

    worksheet.cell(row=rowMaxCount + 1, column=4, value=value)

    wb.save(saveExcelUrl)
    rowMaxCount = rowMaxCount + 1


# 我的代码
def onpressed(Key):
    while True:
        # print(Key)
        if (Key == keyboard.Key.caps_lock):  # 开始
            global start
            start = True
            print('go')

        if (Key == keyboard.Key.f3):  # 结束
            if (saving):
                print('it''s saving !! not yet')
            else:
                sys.exit()
        # print(Key)
        return True


def main():
    while True:
        # 主程序在这
        Do()

def GetExcelUrls(curDirUrl):
    list=[]
    dirlist=os.listdir(curDirUrl)
    for url in dirlist:
        if('.xlsx') in url:
            #list.append(curDirUrl+'\\'+url)
            list.append(url)

    return list
if __name__ == '__main__':
    k = PyKeyboard()
    m = PyMouse()
    scoll_count = 1
    start = False
    saving = False
    curCount = 0
    # saveExcelUrl = r"C:\Users\123\Desktop\广联达\安装\save.xlsx"  # to do-------------
    curExcelUrls=[]#当月的信息价的所有excel
    #updateExcelUrl = r"C:\Users\Administrator\Desktop\Xing.xlsx"  # to do-------------
    updateExcelUrl = r"C:\Users\Administrator\Desktop\Xing.xlsx"  # to do-------------
    curDirUrl=r'C:\Users\Administrator\Desktop\Xing\allxings\1'
    curExcelUrls =GetExcelUrls(curDirUrl)
    print(curExcelUrls)

    threads = []
    t2 = threading.Thread(target=main, args=())
    threads.append(t2)
    for t in threads:
        t.setDaemon(True)
        t.start()
    print('press Capital to start,')
    print('运行完excel点击计算下，不然xing识别不了！！！！！！')
    with keyboard.Listener(on_press=onpressed) as listener:
        listener.join()

